import io, csv
import random
from urllib.parse import quote  # URL-encoding
from math import ceil
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import List, ClassVar, Optional
from pydantic import Field
from decimal import Decimal, ROUND_HALF_UP
from common.services.firebase.firebase_object import FirebaseObject


class TTExport(FirebaseObject):

    ALL_FIELDS: ClassVar[List[str]] = [
        "Campaign ID", "Campaign Status", "Campaign Name", "Advertising Objective",
        "Campaign type", "Sales destination", "Use catalog", "Product source",
        "iOS 14 Dedicated Campaign", "App profile page used", "Campaign Budget Type",
        "Campaign Budget Amount", "Campaign Budget Optimization", "Special ads categories",
        "Realtime API (RTA)", "Ad Group ID", "Ad Group Status", "Ad Group Name",
        "Interaction type", "Shop ads type", "Catalog ID", "Business Center ID of the catalog",
        "TikTok Shop code", "Business Center ID of the TikTok Shop", "Placement Types",
        "Placements", "Include search results", "Block List (Pangle)", "Optimization location",
        "TikTok Pixel ID", "Pixel Event", "App ID", "User Comment", "Video Download",
        "Allow video sharing", "Automated Creative Optimization", "Audience Targeting (VSA)",
        "Include audience ID", "Exclude audience ID", "Retarget audience type",
        "Date range (retarget audience)", "Custom audience", "Smart audience", "Location",
        "Gender", "Age", "Languages", "Spending power", "Household income",
        "Interest Category", "Video interactions", "Creator Category",
        "Creator-related Actions", "Hashtag interactions", "Additional interests",
        "Smart interests & behaviors", "Internet service provider", "Operating System",
        "Connection Type", "Carriers", "Device Price", "OS Versions", "Device model",
        "Category exclusions", "Vertical sensitivity", "Inventory filter",
        "Ad Group Budget Type", "Ad Group Budget Amount", "Start Time", "End Time",
        "Dayparting", "Optimization Goal", "In-App Event", "Secondary Goal",
        "Billing Method", "Click-through window", "View-through window",
        "Engaged view-through window", "Event count", "Frequency Cap", "Bid Strategy",
        "Bid", "Bid for oCPC/M", "Minimum ROAS", "Bid for Secondary Goal", "Delivery Type",
        "Ad ID", "Ad Status", "Ad Name", "Products type", "Product set ID", "Product ID",
        "Use TikTok Account", "Identity Type", "Identity ID",
        "Business Center ID of the identity", "Ad format", "Image Name", "Video Name",
        "Post ID", "Ad Music ID", "Only show as ad", "Catalog video template ID",
        "Destination Page Type (VSA)", "Destination", "Instant page ID", "Interactive add-on ID",
        "Text", "Call to action type", "Call to Action", "Playable ID",
        "Auto Ad - Image Name", "Auto Ad - Video Name", "Auto Ad - Text",
        "Auto ad - call to action type", "Auto Ad - Call to Action",
        "Website type", "Web URL", "Deeplink type", "Deeplink URL",
        "Fallback Type", "Fallback Website URL", "Impression Tracking URL",
        "Click Tracking URL", "TikTok website events", "TikTok app events",
        "TikTok offline events"
    ]
    
    def get_xlsx(self) -> bytes:
        rows = self._build_bulk_rows()

        wb = Workbook()
        ws = wb.active
        ws.title = "TikTok Bulk Upload"

        ws.append(self.ALL_FIELDS)
        for row in rows:
            ws.append([row.get(field, "") for field in self.ALL_FIELDS])

        for col_idx, col_name in enumerate(self.ALL_FIELDS, 1):
            max_len = max(len(str(col_name)), *(len(str(r.get(col_name, ""))) for r in rows)) if rows else len(col_name)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def get_xlsx_from_template(self, sheet_name: Optional[str] = None) -> bytes:
        rows = self._build_bulk_rows()

        wb = load_workbook("./media/exports/tt_template.xlsx", data_only=False)
        ws: Worksheet = self._pick_sheet(wb, "Creogen")

        header_row_idx, col_map = self._find_header_row_and_mapping(ws, self.ALL_FIELDS)
        if header_row_idx is None:
            raise RuntimeError("Не найден ряд заголовков, соответствующий ALL_FIELDS, в шаблоне.")

        start_row = self._detect_first_append_row(ws, header_row_idx, key_col=col_map.get(self.ALL_FIELDS[0]))

        r = start_row
        for row in rows:
            for field, col_idx in col_map.items():
                ws.cell(row=r, column=col_idx, value=row.get(field, ""))
            r += 1

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    # -------------------- helpers --------------------
    
    def _calc_groups(self) -> tuple[list[list[str]], int]:
        files = list(self.file_names or [])
        K = max(1, int(self.ad_creatives_in_adgroup_count))
        if not files:
            return [], 0
        N_actual = ceil(len(files) / K)
        groups: list[list[str]] = []
        for g in range(N_actual):
            start = g * K
            end = start + K
            groups.append(files[start:end])
        return groups, N_actual

    def _pick_sheet(self, wb, sheet_name: Optional[str]) -> Worksheet:
        if sheet_name and sheet_name in wb.sheetnames:
            return wb[sheet_name]
        for name in wb.sheetnames:
            if name.lower().startswith("tiktok") or "upload" in name.lower():
                return wb[name]
        return wb.active

    def _find_header_row_and_mapping(self, ws: Worksheet, fields: List[str]) -> tuple[Optional[int], dict]:
        best_row = None
        best_map = {}
        for row_idx in range(1, min(ws.max_row, 50) + 1):
            values = [str(c.value).strip() if c.value is not None else "" for c in ws[row_idx]]
            if not any(values):
                continue
            m: dict[str, int] = {}
            for col_idx, val in enumerate(values, start=1):
                if val in fields:
                    m[val] = col_idx
            if len(m) > len(best_map):
                best_map, best_row = m, row_idx
                if len(best_map) == len(fields):
                    break
        return best_row, best_map

    def _detect_first_append_row(self, ws: Worksheet, header_row_idx: int, key_col: Optional[int]) -> int:
        if key_col is None:
            return ws.max_row + 1
        first_data_row = header_row_idx + 1
        val = ws.cell(row=first_data_row, column=key_col).value
        if val in (None, ""):
            return first_data_row
        return ws.max_row + 1    
    
    def _even_bids(self, n: int) -> List[float]:
        bmin, bmax = sorted([float(self.bid_min), float(self.bid_max)])
        n = int(n)
        if n <= 0:
            return []
        denom = max(n - 1, 1)
        step = (bmax - bmin) / denom
        return [round(bmin + i * step, 2) for i in range(n)]
    
    def _pick_title_pair(self) -> tuple[str, str]:
        """
        Возвращает (raw_title, encoded_title).
        Если ad_titles пуст — оба пустые.
        """
        if not self.ad_titles:
            return "", ""
        raw = str(random.choice(self.ad_titles))   # равновероятный выбор
        enc = quote(raw, safe="")                  # percent-encoding для URL
        return raw, enc

    def _build_bulk_rows(self) -> List[dict]:
        groups, N_actual = self._calc_groups()
        if N_actual == 0:
            return []

        bids = self._even_bids(N_actual)

        rows: List[dict] = []
        for g, files_in_group in enumerate(groups):
            if not files_in_group:
                continue
            ad_group_name = f"ADG_{g + 1}_{self.id}"
            bid = bids[g]
            for j, vf in enumerate(files_in_group, start=1):
                ad_name = f"AD_{g + 1}_{j}_{self.id}"
                rows.append(
                    self.__generate_row(
                        ad_group_name=ad_group_name,
                        ad_name=ad_name,
                        video_file_name=vf,
                        bid=bid
                    )
                )
        return rows

    def __generate_row(self, ad_group_name: str, ad_name: str, video_file_name: str, bid: float):
        now_str = (datetime.now() - timedelta(days=1)).strftime("%Y/%m/%d %H:%M")

        # Берём титул: сырой и закодированный (один и тот же для строки)
        raw_title, encoded_title = self._pick_title_pair()

        # Подставляем в URL, если есть плейсхолдер
        final_url = self.url
        if isinstance(final_url, str) and "YYYYYYY" in final_url and encoded_title:
            final_url = final_url.replace("YYYYYYY", encoded_title)

        return {
            "Campaign Status": "On",
            "Campaign Name": f"{self.campaign_name}-{self.id}",
            "Advertising Objective": "Sales",
            "Sales destination": "Website",
            "iOS 14 Dedicated Campaign": "Off",
            "Campaign Budget Type": "No Limit",
            "Campaign Budget Optimization": "Off",
            "Ad Group Status": "On",
            "Ad Group Name": ad_group_name,
            "Placement Types": "Select",
            "Placements": "TikTok",
            "Include search results": "Off",
            "Optimization location": "Website",
            "TikTok Pixel ID": self.pixel_id,
            "Pixel Event": self.pixel_event,
            "User Comment": "Off",
            "Video Download": "Off",
            "Allow video sharing": "Off",
            "Automated Creative Optimization": "Off",
            "Smart audience": "Off",
            "Location": ",".join(self.locations),
            "Gender": "All",
            "Age": "18-24,25-34,35-44,45-54,55+",
            "Languages": ",".join(self.languages),
            "Spending power": "All",
            "Household income": "All",
            "Smart interests & behaviors": "Off",
            "Internet service provider": "All",
            "Operating System": "All",
            "Connection Type": "All",
            "Carriers": "All",
            "Device Price": "All",
            "OS Versions": "All",
            "Device model": "All",
            "Inventory filter": "Full inventory",
            "Ad Group Budget Type": "Daily",
            "Ad Group Budget Amount": self._fmt_dot(self.budget),
            "Start Time": now_str,
            "End Time": "No Limit",
            "Dayparting": "All Day",
            "Optimization Goal": "Conversion",
            "Billing Method": "oCPM",
            "Click-through window": "7-day click",
            "View-through window": "1-day view",
            "Event count": "Every",
            "Bid Strategy": "Cost Cap",
            "Bid for oCPC/M": self._fmt_dot(bid),
            "Delivery Type": "Standard",
            "Ad Status": "On",
            "Ad Name": ad_name,
            "Use TikTok Account": "Off",
            "Identity Type": "Custom Identity",
            "Identity ID": self.identity_id,
            "Ad format": "Single video",
            "Video Name": video_file_name,
            "Text": raw_title,
            "Call to action type": "Standard",
            "Call to Action": "Learn More",
            "Playable ID": "",
            "Auto Ad - Image Name": "",
            "Auto Ad - Video Name": "",
            "Auto Ad - Text": "",
            "Auto ad - call to action type": "",
            "Auto Ad - Call to Action": "",
            "Website type": "",
            "Web URL": final_url,
            "Deeplink type": "",
            "Deeplink URL": "",
            "Fallback Type": "",
            "Fallback Website URL": "",
            "Impression Tracking URL": "",
            "Click Tracking URL": "",
            "TikTok website events": self.event_name,
            "TikTok app events": "",
            "TikTok offline events": ""
        }
        
    def _fmt_dot(self, x: float, places: int = 2) -> str:
        q = Decimal(str(x)).quantize(Decimal(10) ** -places, rounding=ROUND_HALF_UP)
        return format(q, f".{places}f")
    
    # -------- поля модели --------
    user_id: Optional[str] = None
    publication_id: Optional[str] = None
    campaign_name: str
    ad_creatives_in_adgroup_count: int
    pixel_id: str
    pixel_event: str
    locations: List[str]
    languages: List[str]
    budget: float
    bid_min: float
    bid_max: float
    identity_id: str
    text: Optional[str] = None  # Deprecated
    url: str
    event_name: str
    file_names: List[str] = Field(default_factory=list)
    ad_titles: List[str] = Field(default_factory=list)

    @staticmethod
    def collection_name():
        return "tt_exports"
